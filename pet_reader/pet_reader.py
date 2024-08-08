from token import NAME
from typing import Any
from openpyxl import load_workbook

from pet_reader.helpers import (
    ABILITY_START,
    BASE_HEALTH,
    BASE_POWER,
    BASE_SPEED,
    BREED,
    COLLECTED,
    FAMILY,
    LEVEL,
    QUALITY,
    row_is_valid,
)
from pets.db import PetDB

from pets.models import Pet, PetSpecies, Stats


class PetReader:
    def __init__(self, db: PetDB):
        self.db = db
        self.file_name = "Pet_Collection_Export_XuFu.xlsx"

    def read_pets(self) -> None:
        collection = self._read_collection()
        species, pets = self._make_pets_from_collection(collection)
        self.db.species = species
        self.db.pets = pets

    def _read_collection(self) -> list[tuple]:
        wb = load_workbook(self.file_name, read_only=True)
        sheet = wb["Collection"]
        out = self._parse_collection_values(
            list(sheet.iter_rows(min_row=2, values_only=True))
        )
        return out

    def _parse_collection_values(
        self, rows: list[tuple[Any, ...]]
    ) -> list[tuple[Any, ...]]:
        return [row for row in rows if row_is_valid(row)]

    def _make_pets_from_collection(
        self, collection: list[tuple]
    ) -> tuple[list[PetSpecies], list[Pet]]:
        species_list = []
        pets = []
        for row in collection:
            species = PetSpecies(
                name=row[NAME],
                family=row[FAMILY],
                base_stats=Stats(
                    health=row[BASE_HEALTH],
                    power=row[BASE_POWER],
                    speed=row[BASE_SPEED],
                ),
                abilities=[x for x in row[ABILITY_START : ABILITY_START + 5]],
            )
            species_list.append(species)
            if row[COLLECTED] == "Yes" and row[LEVEL] == 25 and row[QUALITY] == "Rare":
                pets.append(
                    Pet(
                        species=species,
                        quality=row[QUALITY],
                        breed=row[BREED],
                        level=row[LEVEL],
                    )
                )
        return species_list, pets
